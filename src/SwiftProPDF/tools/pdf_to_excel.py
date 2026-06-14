from pathlib import Path

import fitz
from openpyxl import Workbook

from SwiftProPDF.tools.exceptions import PdfConversionError


def pdf_to_excel(input_path: Path, output_path: Path, overwrite: bool = False) -> None:
    """Convert PDF text content to an Excel workbook."""
    input_path = Path(input_path)
    output_path = Path(output_path)

    if not input_path.exists():
        raise PdfConversionError(f"Input PDF does not exist: {input_path}")

    if output_path.exists() and not overwrite:
        raise PdfConversionError(f"Output file already exists: {output_path}")

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        doc = fitz.open(str(input_path))
        wb = Workbook()
        wb.remove(wb.active)

        for page_idx, page in enumerate(doc):
            text = page.get_text()
            sheet_name = f"Page_{page_idx + 1}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ws = wb.create_sheet(sheet_name)
            lines = text.split("\n")
            for row_idx, line in enumerate(lines, 1):
                cells = line.split() if line.strip() else []
                for col_idx, cell_val in enumerate(cells, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_val)

        doc.close()
        wb.save(str(output_path))

    except Exception as exc:
        raise PdfConversionError(f"Could not convert PDF to Excel: {str(exc)}") from exc
